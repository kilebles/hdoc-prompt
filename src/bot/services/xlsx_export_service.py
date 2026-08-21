from io import BytesIO
from typing import Protocol

from openpyxl import Workbook
from openpyxl.styles import Font

from bot.models.i2v_generation import GeneratedPair


class XlsxExportService(Protocol):
    """Renders generated pairs into a downloadable .xlsx workbook."""

    def build_i2v_workbook(self, pairs: list[GeneratedPair]) -> bytes: ...


class XlsxExportServiceImpl:
    def build_i2v_workbook(self, pairs: list[GeneratedPair]) -> bytes:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Prompts"

        ws["A1"] = "number"
        ws["B1"] = "img_prompt"
        ws["C1"] = "vid_prompt"
        ws["D1"] = "stock_query"
        ws["E1"] = "paragraph"
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row_i, pair in enumerate(pairs, start=2):
            ws[f"A{row_i}"] = f"{pair.paragraph_number}.{pair.pair_number}"
            ws[f"B{row_i}"] = pair.img
            ws[f"C{row_i}"] = pair.vid
            ws[f"D{row_i}"] = pair.stock_query
            ws[f"E{row_i}"] = pair.paragraph_text

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 80

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
